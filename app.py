from flask import Flask, flash, render_template, request, redirect, url_for, send_file, session, jsonify
import psycopg2
from psycopg2 import OperationalError, DatabaseError
from datetime import datetime, date, time
import openpyxl
import re
import os
import csv
import io as _stdio_io
import time as _time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pco_linea2_secret_key_dev")

CV_MAP = {
    'E20': ['2001', '2002'],
    'E21': ['2101', '2102'],
    'E22': ['2203', '2204'],
    'E23': ['2305', '2304'],
    'E24': ['2405', '2402'],
    'PV19': ['PV19'],
    'PV20': ['PV20'],
    'PV21': ['PV21'],
    'PV22': ['PV22'],
    'PV23': ['PV23'],
    'PV24': ['PV24'],
    'E20->E21': ['2001', '2003', '2005', '2101', '2103', '2003a', '2003b', '2101a', '2101b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2102'],
    'E20->E22': ['2001', '2003', '2005', '2101', '2103', '2105', '2107', '2201', '2203', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2102', '2106'],
    'E20->E23': ['2001', '2003', '2005', '2101', '2103', '2105', '2107', '2201', '2203', '2205', '2301', '2303', '2305', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2206', '2302a', '2032b', '2304', '2102', '2106', '2302'],
    'E20->E24': ['2001', '2003', '2005', '2101', '2103', '2105', '2107', '2201', '2203', '2205', '2301', '2303', '2305', '2307', '2401', '2403', '2405', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', '2307a', '2307b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2206', '2302a', '2032b', '2304', '2306', '2308', '2402', '2102', '2106', '2302'],
    'E21->E22': ['2103', '2105', '2107', '2201', '2203', '2105a', '2105b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2106'],
    'E21->E23': ['2103', '2105', '2107', '2201', '2203', '2205', '2301', '2303', '2305', '2105a', '2105b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2206', '2302a', '2032b', '2304', '2106', '2302'],
    'E21->E24': ['2103', '2105', '2107', '2201', '2203', '2205', '2301', '2303', '2305', '2307', '2401', '2403', '2405', '2105a', '2105b', '2307a', '2307b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2206', '2302a', '2032b', '2304', '2306', '2308', '2402', '2106', '2302'],
    'E22->E23': ['2203', '2205', '2301', '2303', '2305', '2204', '2206', '2302a', '2032b', '2304', '2302'],
    'E22->E24': ['2203', '2205', '2301', '2303', '2305', '2307', '2401', '2403', '2405', '2307a', '2307b', '2204', '2206', '2302a', '2032b', '2304', '2306', '2308', '2402', '2302'],
    'E23->E24': ['2305', '2307', '2401', '2403', '2405', '2307a', '2307b', '2304', '2306', '2308', '2402'],
    'E20->PV19': ['2007', '2001', '2010', '2002', 'PV19 BIS'],
    'E20->PV20': ['2001', '2003', '2003a', '2003b', '2002', '2004', '2006', 'PV20'],
    'E20->PV21': ['2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2102', '2106', '2001', '2003', '2005', '2101', '2103', '2105', '2107', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', 'PV21'],
    'E20->PV22': ['2001', '2003', '2005', '2101', '2103', '2105', '2107', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2102', '2106', 'PV21', '2001', '2003', '2005', '2101', '2103', '2105', '2107', '2201', '2203', '2205', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2206', '2102', '2106', 'PV22'],
    'E20->PV23': ['2001', '2003', '2005', '2101', '2103', '2105', '2107', '2201', '2203', '2205', '2301', '2303', '2305', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2206', '2302a', '2032b', '2304', '2102', '2106', '2302', 'PV23'],
    'E20->PV24': ['2001', '2003', '2005', '2101', '2103', '2105', '2107', '2201', '2203', '2205', '2301', '2303', '2305', '2307', '2401', '2403', '2405', '2407', '2409', '2411', '2003a', '2003b', '2101a', '2101b', '2105a', '2105b', '2307a', '2307b', '2002', '2004', '2006', '2008', '2102a', '2102b', '2104', '2106a', '2106b', '2108', '2002', '2204', '2206', '2302a', '2032b', '2304', '2306', '2308', '2402', '2404', '2406', '2102', '2106', '2302', 'PV24'],
    'ZM': ['ZM', '1012'],
    'ZA': ['1011', '1010', '1008b', '1008a', '1006', '1004b', '1004a', '1002', '1008', '1004', '1009', '1007', '1005', '1003', '1001', '1007b', '1007a', '1003b', '1003a', '1009', '1009', '1013', '1013', '1013', '1027', '1027', '1028', '1029', '1029', '1051', 'TK3', '1013', '126', '1030', '1036', '1013', '1013', '1033', '1034', '1035', '1043', 'TK TEST', '1042', '1010', '1031', '1032b', '1032a', '1032'],
    'TK1': ['1012', '1011', '1010', '1008b', '1008a', '1006', '1004b', '1004a', '1002', '1002', '1008', '1004', '1002'],
    'TK2': ['1001', '1009', '1007', '1005', '1003', '1001', '1001', '1007b', '1007a', '1003b', '1003a', '1009', '1009'],
    'TK3': ['1013', '1013', '1027', '1030', '1036'],
    'TK4': ['1013', '1013', '1013', '1033', '1034'],
    'TK5': ['1028'],
    'TK_TEST': ['1032', '1032a', '1032b', '1035', '1031'],
    'TESTTRACK': ['1032', '1032a', '1032b', '1035', '1031'],
}

ZONE_POSITIONS = {
    'D1': [{'top': 88, 'left': 1257, 'width': 132, 'height': 469, 'label': 'D1'}],
    'D2': [{'top': 22, 'left': 1404, 'width': 114, 'height': 547, 'label': 'D2'}],
    'E20': [{'top': 540, 'left': 162, 'width': 31, 'height': 96, 'label': 'E20'}],
    'E20->E21': [
        {'top': 542, 'left': 167, 'width': 403, 'height': 95, 'label': 'E20->E21'},
        {'top': 542, 'left': 562, 'width': 269, 'height': 95, 'label': 'E20->E21'},
    ],
    'E20->E22': [{'top': 542, 'left': 178, 'width': 660, 'height': 93, 'label': 'E20->E22'}],
    'E20->E23': [{'top': 543, 'left': 178, 'width': 909, 'height': 93, 'label': 'E20->E23'}],
    'E20->E24': [{'top': 543, 'left': 174, 'width': 1204, 'height': 93, 'label': 'E20->E24'}],
    'E20->PV19': [
        {'top': 542, 'left': 34, 'width': 141, 'height': 95, 'label': 'E20->PV19'},
        {'top': 505, 'left': 34, 'width': 41, 'height': 11, 'label': 'E20->PV19'},
    ],
    'E20->PV20': [
        {'top': 542, 'left': 165, 'width': 209, 'height': 95, 'label': 'E20->PV20'},
        {'top': 505, 'left': 344, 'width': 31, 'height': 13, 'label': 'E20->PV20'},
        {'top': 543, 'left': 371, 'width': 459, 'height': 95, 'label': 'E20->PV20'},
    ],
    'E20->PV21': [{'top': 542, 'left': 178, 'width': 569, 'height': 95, 'label': 'E20->PV21'}],
    'E20->PV22': [{'top': 542, 'left': 176, 'width': 732, 'height': 94, 'label': 'E20->PV22'}],
    'E20->PV23': [{'top': 542, 'left': 172, 'width': 955, 'height': 94, 'label': 'E20->PV23'}],
    'E20->PV24': [{'top': 542, 'left': 176, 'width': 1420, 'height': 95, 'label': 'E20->PV24'}],
    'E21': [{'top': 539, 'left': 552, 'width': 32, 'height': 99, 'label': 'E21'}],
    'E21->E20': [{'top': 544, 'left': 174, 'width': 393, 'height': 90, 'label': 'E21->E20'}],
    'E21->E22': [{'top': 542, 'left': 566, 'width': 271, 'height': 95, 'label': 'E21->E22'}],
    'E21->E23': [{'top': 543, 'left': 563, 'width': 523, 'height': 95, 'label': 'E21->E23'}],
    'E21->E24': [{'top': 541, 'left': 562, 'width': 816, 'height': 96, 'label': 'E21->E24'}],
    'E21->PV19': [{'top': 542, 'left': 32, 'width': 537, 'height': 95, 'label': 'E21->PV19'}],
    'E21->PV20': [{'top': 543, 'left': 371, 'width': 199, 'height': 95, 'label': 'E21->PV20'}],
    'E21->PV21': [{'top': 542, 'left': 565, 'width': 184, 'height': 95, 'label': 'E21->PV21'}],
    'E21->PV22': [{'top': 543, 'left': 563, 'width': 344, 'height': 95, 'label': 'E21->PV22'}],
    'E21->PV23': [{'top': 541, 'left': 564, 'width': 562, 'height': 97, 'label': 'E21->PV23'}],
    'E21->PV24': [{'top': 541, 'left': 563, 'width': 1031, 'height': 97, 'label': 'E21->PV24'}],
    'E22': [{'top': 540, 'left': 816, 'width': 30, 'height': 95, 'label': 'E22'}],
    'E22->E20': [{'top': 543, 'left': 173, 'width': 657, 'height': 94, 'label': 'E22->E20'}],
    'E22->E23': [{'top': 541, 'left': 829, 'width': 256, 'height': 96, 'label': 'E22->E23'}],
    'E22->E24': [{'top': 543, 'left': 828, 'width': 551, 'height': 92, 'label': 'E22->E24'}],
    'E22->PV19': [{'top': 540, 'left': 32, 'width': 799, 'height': 97, 'label': 'E22->PV19'}],
    'E22->PV21': [{'top': 542, 'left': 745, 'width': 85, 'height': 97, 'label': 'E22->PV21'}],
    'E22->PV22': [{'top': 543, 'left': 829, 'width': 79, 'height': 93, 'label': 'E22->PV22'}],
    'E22->PV24': [{'top': 543, 'left': 827, 'width': 768, 'height': 93, 'label': 'E22->PV24'}],
    'E22-PV23': [{'top': 542, 'left': 831, 'width': 295, 'height': 95, 'label': 'E22-PV23'}],
    'E23': [{'top': 540, 'left': 1065, 'width': 30, 'height': 98, 'label': 'E23'}],
    'E23->E20': [{'top': 542, 'left': 173, 'width': 907, 'height': 96, 'label': 'E23->E20'}],
    'E23->E21': [{'top': 543, 'left': 563, 'width': 517, 'height': 97, 'label': 'E23->E21'}],
    'E23->E22': [{'top': 542, 'left': 826, 'width': 255, 'height': 97, 'label': 'E23->E22'}],
    'E23->E24': [{'top': 541, 'left': 1079, 'width': 299, 'height': 95, 'label': 'E23->E24'}],
    'E23->PV19': [{'top': 542, 'left': 33, 'width': 1047, 'height': 95, 'label': 'E23->PV19'}],
    'E23->PV20': [{'top': 542, 'left': 371, 'width': 708, 'height': 95, 'label': 'E23->PV20'}],
    'E23->PV21': [{'top': 542, 'left': 747, 'width': 334, 'height': 94, 'label': 'E23->PV21'}],
    'E23->PV22': [{'top': 543, 'left': 904, 'width': 177, 'height': 95, 'label': 'E23->PV22'}],
    'E23->PV23': [{'top': 541, 'left': 1078, 'width': 48, 'height': 94, 'label': 'E23->PV23'}],
    'E23->PV24': [{'top': 542, 'left': 1078, 'width': 518, 'height': 95, 'label': 'E23->PV24'}],
    'E24': [{'top': 540, 'left': 1350, 'width': 45, 'height': 98, 'label': 'E24'}],
    'E24->E20': [{'top': 542, 'left': 173, 'width': 1199, 'height': 95, 'label': 'E24->E20'}],
    'E24->E21': [{'top': 541, 'left': 564, 'width': 812, 'height': 96, 'label': 'E24->E21'}],
    'E24->E22': [{'top': 542, 'left': 826, 'width': 546, 'height': 96, 'label': 'E24->E22'}],
    'E24->E23': [{'top': 543, 'left': 1076, 'width': 301, 'height': 93, 'label': 'E24->E23'}],
    'E24->PV19': [{'top': 542, 'left': 31, 'width': 1343, 'height': 96, 'label': 'E24->PV19'}],
    'E24->PV20': [{'top': 542, 'left': 370, 'width': 1005, 'height': 96, 'label': 'E24->PV20'}],
    'E24->PV21': [{'top': 542, 'left': 745, 'width': 630, 'height': 95, 'label': 'E24->PV21'}],
    'E24->PV22': [{'top': 543, 'left': 906, 'width': 464, 'height': 96, 'label': 'E24->PV22'}],
    'E24->PV23': [{'top': 542, 'left': 1123, 'width': 254, 'height': 93, 'label': 'E24->PV23'}],
    'E24->PV24': [{'top': 543, 'left': 1367, 'width': 228, 'height': 94, 'label': 'E24->PV24'}],
    'PV19': [{'top': 502, 'left': 33, 'width': 44, 'height': 16, 'label': 'PV19'}],
    'PV20': [{'top': 503, 'left': 343, 'width': 31, 'height': 16, 'label': 'PV20'}],
    'PV21': [{'top': 501, 'left': 718, 'width': 32, 'height': 17, 'label': 'PV21'}],
    'PV22': [{'top': 501, 'left': 877, 'width': 29, 'height': 18, 'label': 'PV22'}],
    'PV23': [{'top': 500, 'left': 1092, 'width': 32, 'height': 20, 'label': 'PV23'}],
    'PV24': [{'top': 503, 'left': 1570, 'width': 26, 'height': 16, 'label': 'PV24'}],
    'TK_TEST': [{'top': 442, 'left': 188, 'width': 911, 'height': 40, 'label': 'TK TEST'}],
    'TK1': [
        {'top': 89, 'left': 291, 'width': 1098, 'height': 36, 'label': 'TK1'},
        {'top': 95, 'left': 293, 'width': 47, 'height': 166, 'label': 'TK1'},
        {'top': 87, 'left': 1324, 'width': 62, 'height': 351, 'label': 'TK1'},
    ],
    'TK2': [
        {'top': 22, 'left': 160, 'width': 74, 'height': 164, 'label': 'TK2'},
        {'top': 20, 'left': 162, 'width': 1324, 'height': 41, 'label': 'TK2'},
        {'top': 20, 'left': 1407, 'width': 77, 'height': 415, 'label': 'TK2'},
    ],
    'TK3': [
        {'top': 269, 'left': 233, 'width': 223, 'height': 47, 'label': 'TK3'},
        {'top': 303, 'left': 431, 'width': 54, 'height': 68, 'label': 'TK3'},
        {'top': 333, 'left': 487, 'width': 612, 'height': 36, 'label': 'TK3'},
    ],
    'TK4': [{'top': 377, 'left': 277, 'width': 827, 'height': 50, 'label': 'TK4'}],
    'TK5': [{'top': 282, 'left': 456, 'width': 390, 'height': 35, 'label': 'TK5'}],
    'ZM': [
        {'top': 225, 'left': 452, 'width': 297, 'height': 37, 'label': 'ZM'},
        {'top': 198, 'left': 547, 'width': 71, 'height': 36, 'label': 'ZM'},
    ],
    'TK7': [{'top': 442, 'left': 188, 'width': 911, 'height': 40, 'label': 'TK TEST'}],
    'PTSA': [
        {'top': 89, 'left': 291, 'width': 1098, 'height': 36, 'label': 'PTSA'},
        {'top': 22, 'left': 160, 'width': 74, 'height': 164, 'label': 'PTSA'},
        {'top': 20, 'left': 162, 'width': 1324, 'height': 41, 'label': 'PTSA'},
    ],
    'TKTEST': [{'top': 442, 'left': 188, 'width': 911, 'height': 40, 'label': 'TK TEST'}],
    'TESTTRACK': [{'top': 442, 'left': 188, 'width': 911, 'height': 40, 'label': 'TK TEST'}],
}


def get_rects(zona):
    entry = ZONE_POSITIONS.get(zona)
    if entry is None:
        return []
    if isinstance(entry, list):
        return entry
    return [entry]


def get_cvs_for_zonas(ubicacion_zona_str):
    partes = normalizar_zonas(ubicacion_zona_str)
    cvs_encontrados = set()
    for parte in partes:
        if parte in CV_MAP:
            for cv in CV_MAP[parte]:
                cvs_encontrados.add(cv)
    return sorted(cvs_encontrados)


def get_zonas_con_info(ubicacion_zona_str):
    partes = normalizar_zonas(ubicacion_zona_str)
    resultado = []
    for parte in partes:
        rects = get_rects(parte)
        cv_list = CV_MAP.get(parte, [])
        resultado.append({
            'nombre': parte,
            'rects': rects,
            'cvs': cv_list
        })
    return resultado

def get_db_connection(max_retries=3, delay=1):
    database_url = os.environ.get("DATABASE_URL")
    last_err = None
    for i in range(max_retries):
        try:
            if database_url:
                conn = psycopg2.connect(database_url, connect_timeout=10)
            else:
                conn = psycopg2.connect(dbname="pco_db", user="ardalax", host="localhost", port="5432")
            # Fijar la sesion a hora de Lima para que CURRENT_TIME/CURRENT_DATE
            # reflejen la fecha/hora local peruana (UTC-5, sin horario de verano).
            try:
                cur = conn.cursor()
                cur.execute("SET TIME ZONE 'America/Lima';")
                cur.close()
            except Exception:
                pass
            return conn
        except OperationalError as e:
            last_err = e
            if i < max_retries - 1:
                _time.sleep(delay * (i + 1))
            continue
    raise last_err


def normalizar_zonas(zonas):
    if not zonas:
        return set()

    if isinstance(zonas, str):
        partes = zonas.replace(',', '+').split('+')
    else:
        partes = zonas

    zonas_normalizadas = set()
    for parte in partes:
        zona = parte.strip().replace(" ", "").upper()
        if zona:
            zonas_normalizadas.add(zona)
    return zonas_normalizadas


def evaluar_safety_ingreso(ubicacion_zona, trabajos_activos, usa_vehiculo, tipo_vehiculo, codigo_vehiculo, empresa):
    partes_nuevas = normalizar_zonas(ubicacion_zona)

    for trab in trabajos_activos:
        empresa_activa = trab[0]
        zona_activa_cadena = trab[1]
        en_via_usa_vehiculo = bool(trab[2])
        en_via_tipo_vehiculo = trab[3] or "Vehículo Auxiliar"
        en_via_codigo_vehiculo = trab[4] or "sin código"

        partes_activas = normalizar_zonas(zona_activa_cadena)
        coincidencias = partes_nuevas.intersection(partes_activas)

        if not coincidencias:
            continue

        if en_via_usa_vehiculo or usa_vehiculo:
            if en_via_usa_vehiculo and usa_vehiculo:
                mensaje = f"❌ RECHAZADO POR SAFETY: La zona '{', '.join(sorted(coincidencias))}' ya tiene movimiento de {en_via_tipo_vehiculo} ({en_via_codigo_vehiculo}) por la empresa '{empresa_activa}' y no se permite otro vehículo en la misma zona."
            elif en_via_usa_vehiculo:
                mensaje = f"❌ RECHAZADO POR SAFETY: La zona '{', '.join(sorted(coincidencias))}' ya tiene movimiento de {en_via_tipo_vehiculo} ({en_via_codigo_vehiculo}) por la empresa '{empresa_activa}'."
            else:
                vehiculo_label = tipo_vehiculo or "vehículo auxiliar"
                codigo_label = codigo_vehiculo or "sin código"
                mensaje = f"❌ RECHAZADO POR SAFETY: No puedes ingresar el {vehiculo_label} ({codigo_label}) a la zona '{', '.join(sorted(coincidencias))}' porque actualmente hay una cuadrilla peatonal de '{empresa_activa}' trabajando ahí."
            return True, mensaje

    return False, ""

@app.route('/')
def index():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, turno, empresa, orden_trabajo, responsable, ubicacion_zona, 
               num_personas, vaf_tren, conductor, tetra, hora_inicio, hora_fin, 
               operador_turno, spco_turno, estado, 
               usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo, comentario
        FROM seguimiento_vias 
        WHERE archivado = FALSE AND (fecha = CURRENT_DATE OR estado = 'En Vía')
        ORDER BY (hora_fin IS NOT NULL) ASC, hora_inicio DESC;
    """)
    registros = cur.fetchall()
    cur.close()
    conn.close()
    
    turno_actual = session.get('turno', 'Turno 1')
    operador_actual = session.get('operador_turno', '')
    spco_actual = session.get('spco_turno', '')
    
    # LA MAGIA PARA QUE NO SE BORRE: Leemos y limpiamos la memoria al mismo tiempo
    form_previo = session.pop('form_previo', {})
    
    importados_pendientes = session.get('importados_pendientes', [])
    
    return render_template('index.html', 
                           registros=registros, 
                           turno_actual=turno_actual, 
                           operador_actual=operador_actual, 
                           spco_actual=spco_actual,
                           form_previo=form_previo,
                           importados_pendientes=importados_pendientes)

@app.route('/mapa')
def mapa():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT ubicacion_zona, empresa, usa_vehiculo, tipo_vehiculo, codigo_vehiculo, responsable, num_personas, tetra FROM seguimiento_vias WHERE hora_fin IS NULL;")
    registros_activos = cur.fetchall()
    cur.close()
    conn.close()

    zonas_ocupadas = []
    zonas_data = []
    trabajos_resumen = []
    for reg in registros_activos:
        cadena_zona = reg[0]
        empresa = reg[1] or ''
        usa_vehiculo = bool(reg[2])
        tipo_vehiculo = reg[3] or ''
        codigo_vehiculo = reg[4] or ''
        responsable = reg[5] or ''
        num_personas = reg[6] or 0
        tetra = reg[7] or ''

        trabajos_resumen.append({
            'zona': cadena_zona,
            'empresa': empresa,
            'responsable': responsable,
            'num_personas': num_personas,
            'tetra': tetra,
            'usa_vehiculo': usa_vehiculo,
            'tipo_vehiculo': tipo_vehiculo,
            'codigo_vehiculo': codigo_vehiculo,
        })

        partes = cadena_zona.replace(',', '+').split('+')
        for parte in partes:
            zona_limpia = parte.strip().replace(" ", "").upper()
            if zona_limpia and zona_limpia not in zonas_ocupadas:
                zonas_ocupadas.append(zona_limpia)
            if zona_limpia:
                rects = get_rects(zona_limpia)
                if rects:
                    for rect in rects:
                        zonas_data.append({
                            'nombre': zona_limpia,
                            'top': rect.get('top', 0),
                            'left': rect.get('left', 0),
                            'width': rect.get('width', 100),
                            'height': rect.get('height', 50),
                            'label': rect.get('label', zona_limpia),
                            'empresa': empresa,
                            'usa_vehiculo': usa_vehiculo,
                            'tipo_vehiculo': tipo_vehiculo,
                            'codigo_vehiculo': codigo_vehiculo,
                            'responsable': responsable,
                            'num_personas': num_personas,
                            'tetra': tetra,
                        })
                else:
                    zonas_data.append({
                        'nombre': zona_limpia,
                        'top': 0, 'left': 0, 'width': 100, 'height': 50,
                        'label': zona_limpia,
                        'empresa': empresa,
                        'usa_vehiculo': usa_vehiculo,
                        'tipo_vehiculo': tipo_vehiculo,
                        'codigo_vehiculo': codigo_vehiculo,
                        'responsable': responsable,
                        'num_personas': num_personas,
                        'tetra': tetra,
                    })

    return render_template('mapa.html', zonas_ocupadas=zonas_ocupadas, zonas_data=zonas_data, trabajos_resumen=trabajos_resumen)

@app.route('/coordenadas')
def coordenadas():
    return render_template('coordinate_finder.html')

@app.route('/configurar_turno', methods=['POST'])
def configurar_turno():
    session['turno'] = request.form['turno']
    session['operador_turno'] = request.form['operador_turno']
    session['spco_turno'] = request.form['spco_turno']
    return redirect(url_for('index'))

@app.route('/verificar_safety', methods=['POST'])
def verificar_safety():
    data = request.get_json() or {}
    ubicacion_zona = data.get('ubicacion_zona', '')
    usa_vehiculo = True if data.get('usa_vehiculo') == 'si' or data.get('usa_vehiculo') is True else False
    tipo_vehiculo = data.get('tipo_vehiculo', '')
    codigo_vehiculo = data.get('codigo_vehiculo', '')
    empresa = data.get('empresa', '')

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT empresa, ubicacion_zona, usa_vehiculo, tipo_vehiculo, codigo_vehiculo 
        FROM seguimiento_vias 
        WHERE hora_fin IS NULL;
    """)
    trabajos_activos = cur.fetchall()
    cur.close()
    conn.close()

    bloqueo, mensaje = evaluar_safety_ingreso(
        ubicacion_zona,
        trabajos_activos,
        usa_vehiculo,
        tipo_vehiculo,
        codigo_vehiculo,
        empresa
    )

    return jsonify({
        "bloqueo": bloqueo,
        "mensaje": mensaje
    })

@app.route('/ingresar', methods=['POST'])
def ingresar():
    if request.method == 'POST':
        turno = session.get('turno', 'Turno 1')
        operador = session.get('operador_turno', 'No Registrado')
        spco = session.get('spco_turno', 'No Registrado')
        
        empresa = request.form['empresa']
        orden_trabajo = request.form.get('orden_trabajo', '-')
        responsable = request.form['responsable']
        ubicacion_zona = request.form['ubicacion_zona']
        num_personas = request.form['num_personas'] or 0
        tetra = request.form['tetra']
        comentario = request.form.get('comentario', '').strip()

        usa_vehiculo = True if request.form.get('usa_vehiculo') == 'si' else False
        tipo_vehiculo = request.form.get('tipo_vehiculo', '')
        codigo_vehiculo = request.form.get('codigo_vehiculo', '')
        conductor_vehiculo = request.form.get('conductor_vehiculo', '')

        # GUARDAMOS TODO EN LA SESIÓN (Por si hay rechazo de Safety)
        session['form_previo'] = request.form.to_dict()

        partes_nuevas = [p.strip().replace(" ", "").upper() for p in ubicacion_zona.replace(',', '+').split('+') if p.strip()]

        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            SELECT empresa, ubicacion_zona, usa_vehiculo, tipo_vehiculo, codigo_vehiculo 
            FROM seguimiento_vias 
            WHERE hora_fin IS NULL;
        """)
        trabajos_activos = cur.fetchall()
        
        # --- MOTOR DE SAFETY ---
        bloqueo_seguridad, mensaje_alerta = evaluar_safety_ingreso(
            ubicacion_zona,
            trabajos_activos,
            usa_vehiculo,
            tipo_vehiculo,
            codigo_vehiculo,
            empresa
        )

        # --- SI HAY BLOQUEO, REDIRIGIMOS PARA MOSTRAR ALERTA ---
        if bloqueo_seguridad:
            cur.close()
            conn.close()
            flash(mensaje_alerta, "danger")
            return redirect(url_for('index'))

        # --- SI NO HAY PELIGRO, SE INSERTA NORMAL ---
        cur.execute("""
            INSERT INTO seguimiento_vias 
            (turno, operador_turno, spco_turno, empresa, orden_trabajo, responsable, ubicacion_zona, num_personas, tetra, fecha, hora_inicio, estado, usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo, comentario)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_DATE, CURRENT_TIME, 'En Vía', %s, %s, %s, %s, %s);
        """, (turno, operador, spco, empresa, orden_trabajo, responsable, ubicacion_zona, num_personas, tetra, usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo, comentario))
        conn.commit()
        cur.close()
        conn.close()
        
        # Como fue exitoso, borramos la memoria para dejar el formulario limpio
        session.pop('form_previo', None)
        return redirect(url_for('index'))
    
@app.route('/liberar/<int:id>')
def liberar(id):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE seguimiento_vias SET hora_fin = CURRENT_TIME, estado = 'Liberado' WHERE id = %s;", (id,))
        conn.commit()
        flash('Registro liberado correctamente.', 'success')
    except (OperationalError, DatabaseError) as e:
        flash('Error de base de datos al liberar. Intenta de nuevo.', 'danger')
        app.logger.error(f"liberar DB error: {e}")
        if conn:
            conn.rollback()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()
    return redirect(url_for('index'))

@app.route('/revertir/<int:id>')
def revertir(id):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE seguimiento_vias SET hora_fin = NULL, estado = 'En Vía' WHERE id = %s;", (id,))
        conn.commit()
        flash('Registro revertido correctamente.', 'success')
    except (OperationalError, DatabaseError) as e:
        flash('Error de base de datos al revertir. Intenta de nuevo.', 'danger')
        app.logger.error(f"revertir DB error: {e}")
        if conn:
            conn.rollback()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()
    return redirect(url_for('index'))

@app.route('/eliminar/<int:id>')
def eliminar(id):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM seguimiento_vias WHERE id = %s;", (id,))
        conn.commit()
        flash('Registro eliminado correctamente.', 'success')
    except (OperationalError, DatabaseError) as e:
        flash('Error de base de datos al eliminar. Intenta de nuevo.', 'danger')
        app.logger.error(f"eliminar DB error: {e}")
        if conn:
            conn.rollback()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()
    return redirect(url_for('index'))

@app.route('/editar/<int:id>', methods=['POST'])
def editar(id):
    if request.method == 'POST':
        empresa = request.form['empresa']
        orden_trabajo = request.form.get('orden_trabajo', '-')
        responsable = request.form['responsable']
        ubicacion_zona = request.form['ubicacion_zona']
        num_personas = request.form['num_personas'] or 0
        tetra = request.form['tetra']
        comentario = request.form.get('comentario', '').strip()
        
        usa_vehiculo = True if request.form.get('usa_vehiculo') == 'si' else False
        tipo_vehiculo = request.form.get('tipo_vehiculo', '') if usa_vehiculo else ''
        codigo_vehiculo = request.form.get('codigo_vehiculo', '') if usa_vehiculo else ''
        conductor_vehiculo = request.form.get('conductor_vehiculo', '') if usa_vehiculo else ''
        
        h_inicio_str = request.form['hora_inicio']
        h_fin_str = request.form['hora_fin']
        hora_inicio = datetime.strptime(h_inicio_str, "%H:%M:%S").time() if h_inicio_str else None
        hora_fin = datetime.strptime(h_fin_str, "%H:%M:%S").time() if h_fin_str else None
        estado = 'Liberado' if hora_fin else 'En Vía'

        conn = None
        cur = None
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("""
                UPDATE seguimiento_vias 
                SET empresa = %s, orden_trabajo = %s, responsable = %s, ubicacion_zona = %s, num_personas = %s, 
                    tetra = %s, hora_inicio = %s, hora_fin = %s, estado = %s,
                    usa_vehiculo = %s, tipo_vehiculo = %s, codigo_vehiculo = %s, conductor_vehiculo = %s,
                    comentario = %s
                WHERE id = %s;
            """, (empresa, orden_trabajo, responsable, ubicacion_zona, num_personas, tetra, hora_inicio, hora_fin, estado, usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo, comentario, id))
            conn.commit()
            flash('Registro actualizado correctamente.', 'success')
        except (OperationalError, DatabaseError) as e:
            flash('Error de base de datos al actualizar. Intenta de nuevo.', 'danger')
            app.logger.error(f"editar DB error: {e}")
            if conn:
                conn.rollback()
        finally:
            if cur:
                cur.close()
            if conn:
                conn.close()
        return redirect(url_for('index'))

@app.route('/exportar')
def exportar():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT fecha, turno, operador_turno, spco_turno, empresa, responsable, ubicacion_zona, 
               num_personas, tetra, hora_inicio, hora_fin, estado,
               usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo 
        FROM seguimiento_vias 
        WHERE fecha = CURRENT_DATE 
        ORDER BY hora_inicio ASC;
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seguimiento PCO"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells('A1:P1')
    ws['A1'] = "REPORTE DIARIO DE SEGUIMIENTO DE TRABAJOS EN VÍA"
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    headers = ["Fecha", "Turno", "Operador ATS", "Supervisor SPCO", "Empresa", "Responsable", "Zona Ocupada", 
               "N° Pers.", "TETRA", "Hora Inicio", "Hora Fin", "Estado", "¿Usa Vehículo?", "Tipo Vehículo", "Código Vehículo", "Conductor"]
    ws.append([]) 
    ws.append(headers)

    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='000000')
    thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'), top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(value, (datetime, date)):
                cell.value = value.strftime('%Y-%m-%d')
            elif hasattr(value, 'strftime'): 
                cell.value = value.strftime('%H:%M:%S')
            elif isinstance(value, bool):
                cell.value = "Sí" if value else "No"
            else:
                cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r_idx].height = 20

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    fecha_str = datetime.now().strftime('%Y%m%d')
    return send_file(excel_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'Seguimiento_Via_{fecha_str}.xlsx')

@app.route('/historial')
def historial():
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    empresa = request.args.get('empresa', '')
    buscar_texto = request.args.get('buscar_texto', '')
    hora_inicio_desde = request.args.get('hora_inicio_desde', '')
    hora_inicio_hasta = request.args.get('hora_inicio_hasta', '')
    bloque_horario = request.args.get('bloque_horario', '')

    # Resolucion de bloque horario a rango de horas
    if bloque_horario == 'dia':
        hora_inicio_desde, hora_inicio_hasta = '05:00', '17:00'
    elif bloque_horario == 'noche':
        hora_inicio_desde, hora_inicio_hasta = '17:00', '05:00'

    query = """
        SELECT id, fecha, turno, empresa, orden_trabajo, responsable, ubicacion_zona, 
               num_personas, tetra, hora_inicio, hora_fin, estado, 
               usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo,
               operador_turno, spco_turno, comentario
        FROM seguimiento_vias 
        WHERE 1=1
    """
    params = []

    if fecha_inicio:
        query += " AND fecha >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND fecha <= %s"
        params.append(fecha_fin)
    if empresa:
        query += " AND empresa = %s"
        params.append(empresa)
    if buscar_texto:
        query += " AND (responsable ILIKE %s OR ubicacion_zona ILIKE %s OR codigo_vehiculo ILIKE %s OR conductor_vehiculo ILIKE %s)"
        term = f"%{buscar_texto}%"
        params.extend([term, term, term, term])
    if hora_inicio_desde and hora_inicio_hasta and hora_inicio_desde <= hora_inicio_hasta:
        # Rango simple dentro del mismo dia: 05:00 a 17:00
        query += " AND hora_inicio >= %s AND hora_inicio <= %s"
        params.append(hora_inicio_desde)
        params.append(hora_inicio_hasta)
    elif hora_inicio_desde and hora_inicio_hasta and hora_inicio_desde > hora_inicio_hasta:
        # Rango que cruza medianoche: 17:00 a 05:00 del dia siguiente
        query += " AND (hora_inicio >= %s OR hora_inicio <= %s)"
        params.append(hora_inicio_desde)
        params.append(hora_inicio_hasta)

    query += " ORDER BY fecha DESC, hora_inicio DESC;"

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, tuple(params))
    registros = cur.fetchall()

    cur.execute("SELECT DISTINCT empresa FROM seguimiento_vias WHERE empresa IS NOT NULL ORDER BY empresa;")
    lista_empresas = [row[0] for row in cur.fetchall()]
    
    cur.close()
    conn.close()

    return render_template('historial.html', 
                           registros=registros, 
                           lista_empresas=lista_empresas,
                           fecha_inicio=fecha_inicio,
                           fecha_fin=fecha_fin,
                           empresa_seleccionada=empresa,
                           buscar_texto=buscar_texto,
                           hora_inicio_desde=hora_inicio_desde,
                           hora_inicio_hasta=hora_inicio_hasta,
                           bloque_horario=bloque_horario)

@app.route('/exportar_historial')
def exportar_historial():
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    empresa = request.args.get('empresa', '')
    buscar_texto = request.args.get('buscar_texto', '')

    query = """
        SELECT fecha, turno, operador_turno, spco_turno, empresa, responsable, ubicacion_zona, 
               num_personas, tetra, hora_inicio, hora_fin, estado,
               usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo, comentario 
        FROM seguimiento_vias 
        WHERE 1=1
    """
    params = []

    if fecha_inicio:
        query += " AND fecha >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND fecha <= %s"
        params.append(fecha_fin)
    if empresa:
        query += " AND empresa = %s"
        params.append(empresa)
    if buscar_texto:
        query += " AND (responsable ILIKE %s OR ubicacion_zona ILIKE %s OR codigo_vehiculo ILIKE %s OR conductor_vehiculo ILIKE %s OR comentario ILIKE %s)"
        term = f"%{buscar_texto}%"
        params.extend([term, term, term, term, term])

    query += " ORDER BY fecha ASC, hora_inicio ASC;"

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Filtrado PCO"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells('A1:Q1')
    ws['A1'] = "REPORTE HISTÓRICO DE SEGUIMIENTO DE TRABAJOS EN VÍA"
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='333f48', end_color='333f48', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    headers = ["Fecha", "Turno", "Operador ATS", "Supervisor SPCO", "Empresa", "Responsable", "Zona Ocupada", 
               "N° Pers.", "TETRA", "Hora Inicio", "Hora Fin", "Estado", "¿Usa Vehículo?", "Tipo Vehículo", "Código Vehículo", "Conductor", "Comentario"]
    ws.append([]) 
    ws.append(headers)

    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='000000')
    thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'), top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(value, (datetime, date)):
                cell.value = value.strftime('%Y-%m-%d')
            elif hasattr(value, 'strftime'): 
                cell.value = value.strftime('%H:%M:%S')
            elif isinstance(value, bool):
                cell.value = "Sí" if value else "No"
            else:
                cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            # Alinear al centros; comentarios a la izquierda con wrap para que se vea completo
            if c_idx == 17:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r_idx].height = 20

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        # comentarios: ancho mínimo mayor y máximo de 60 (wrap activo)
        if col_letter == 'Q':
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 60), 25)
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return send_file(excel_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Historial_Reporte_Filtrado.xlsx')

@app.route('/archivar_turno', methods=['POST'])
def archivar_turno():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE seguimiento_vias SET archivado = TRUE WHERE estado = 'Liberado' AND archivado = FALSE;")
    conn.commit()
    cur.close()
    conn.close()
    
    flash("✅ Relevo Preparado: Los trabajos liberados han sido archivados. Los trabajos 'En Vía' se mantienen en pantalla.", "success")
    return redirect(url_for('index'))

@app.route('/restaurar/<int:id>')
def restaurar(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE seguimiento_vias SET archivado = FALSE, fecha = CURRENT_DATE WHERE id = %s;", (id,))
    conn.commit()
    cur.close()
    conn.close()
    
    flash("✅ Registro restaurado exitosamente. El trabajo está de vuelta en tu panel de Monitoreo para que puedas corregirlo.", "success")
    return redirect(url_for('index'))

def extract_tetra(texto):
    if not texto:
        return ''
    texto = str(texto).replace(' ', '').replace('\n', ' ')
    matches = re.findall(r'(?<!\d)22\d{3}(?!\d)', texto)
    return ', '.join(matches) if matches else ''


@app.route('/importar_excel', methods=['POST'])
def importar_excel():
    if 'excel_file' not in request.files:
        return jsonify({'error': 'No se subió archivo'}), 400
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # ====== DETECCIÓN AUTOMÁTICA DE CABECERAS (fila con "ESTADO") ======
    # Busca la fila que contiene la cabecera y mapea columnas dinámicamente,
    # para soportar cualquier variación del orden de columnas del Excel origen.
    def _norm_txt(s):
        if s is None:
            return ''
        if not isinstance(s, str):
            s = str(s)
        s = s.strip().upper()
        s = s.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U').replace('Ñ', 'N')
        s = re.sub(r'\s+', ' ', s)
        return s

    CABECERAS_XLSX = {
        'ESTADO': 'estado',
        'SPCO QUE VALIDA': 'spco',
        'SEGUIMIENTO (NOMBRE OPCO)': 'operador',
        'SEGUIMIENTO (NOMBRE OPC)': 'operador',
        'ORDEN DE TRABAJO': 'orden_trabajo',
        'EMPRESA CONTRATISTA': 'empresa',
        'ACCESO A VÍA': 'acceso',
        'ACCESO A VIA': 'acceso',
        'DESDE': 'desde',
        'HASTA': 'hasta',
        'RESPONSABLE DE TRABAJOS': 'responsable',
        'CELULAR Y TETRA EXT': 'celular_tetra',
        # Comentario del permiso: dos posibles cabeceras (en orden de preferencia)
        'REQUERIMIENTOS ADICIONALES / COMENTARIOS': 'comentario',
        'REQUERIMIENTOS ADICIONALE': 'comentario',
        'REQUERIMENTOS ADICIONALES / COMENTARIOS': 'comentario',
        'COMENTARIOS DEL PCO': 'comentario_pco',
    }

    fila_cabecera = None
    col_idx = {}
    max_filas_buscar = min(ws.max_row, 15)
    for r in range(1, max_filas_buscar + 1):
        valores_normalizados = [_norm_txt(ws.cell(row=r, column=c).value) for c in range(1, 32)]
        if 'ESTADO' in valores_normalizados or 'ESTADO ' in [v + ' ' for v in valores_normalizados]:
            for c_idx, val in enumerate(valores_normalizados, 1):
                if val in CABECERAS_XLSX:
                    campo = CABECERAS_XLSX[val]
                    # columna 23 ("REQUERIMIENTOS") tiene prioridad sobre columna 8 ("COMENTARIOS PCO")
                    if campo == 'comentario':
                        if 'comentario' not in col_idx:
                            col_idx['comentario'] = c_idx
                    elif campo == 'comentario_pco':
                        if 'comentario_pco' not in col_idx:
                            col_idx['comentario_pco'] = c_idx
                    else:
                        col_idx[campo] = c_idx
            fila_cabecera = r
            break

    if fila_cabecera is None:
        # Fallback al mapeo posicional histórico (companibilidad hacia atras)
        fila_cabecera = 5
        col_idx = {
            'estado': 2, 'spco': 3, 'operador': 5, 'orden_trabajo': 10,
            'empresa': 15, 'acceso': 17, 'desde': 21, 'hasta': 22,
            'responsable': 26, 'celular_tetra': 27,
            'comentario': 23, 'comentario_pco': 8,
        }

    fila_inicio_datos = fila_cabecera + 1

    datos = []
    for row in range(fila_inicio_datos, ws.max_row + 1):
        def _get(campo):
            idx = col_idx.get(campo)
            if idx is None or idx > ws.max_column:
                return ''
            v = ws.cell(row=row, column=idx).value
            return v if v is not None else ''

        estado_val = _get('estado')
        acceso_val = _get('acceso')

        if str(estado_val).strip().upper() != 'CONFIRMADA':
            continue
        if str(acceso_val).strip().upper() != 'SI':
            continue

        spco = _get('spco')
        operador = _get('operador')
        desde = _get('desde')
        hasta = _get('hasta')
        responsable = _get('responsable')
        celular_tetra = _get('celular_tetra')
        orden_trabajo = _get('orden_trabajo')
        empresa = _get('empresa')

        tetra = extract_tetra(celular_tetra)

        # Comentario del permiso de trabajo:
        # Prioridad 1: "REQUERIMIENTOS ADICIONALES / COMENTARIOS" (col 23, más descriptivo)
        # Prioridad 2: "COMENTARIOS DEL PCO" (col 8, suele venir vacío)
        # Si ambos existen y el principal está vacío, usar el secundario.
        comentario_permiso = ''
        if 'comentario' in col_idx:
            comentario_permiso = str(_get('comentario')).strip()
        if not comentario_permiso and 'comentario_pco' in col_idx:
            comentario_permiso = str(_get('comentario_pco')).strip()
        # Limpieza: si solo trae un nbsp o类似的
        if comentario_permiso in ('\\xa0', '\xa0', 'N/A', 'n/a'):
            comentario_permiso = ''

        datos.append({
            'spco': str(spco).strip(),
            'operador': str(operador).strip(),
            'desde': str(desde).strip(),
            'hasta': str(hasta).strip(),
            'responsable': str(responsable).strip(),
            'tetra': tetra,
            'orden_trabajo': str(orden_trabajo).strip(),
            'empresa': str(empresa).strip(),
            'comentario': comentario_permiso,
        })

    wb.close()

    # Guardar en sesión para persistir entre recargas
    prev = session.get('importados_pendientes', [])
    prev.extend(datos)
    session['importados_pendientes'] = prev

    return jsonify({'datos': datos})


@app.route('/importar_texto', methods=['POST'])
def importar_texto():
    """Importa trabajos pegados como texto TSV (mismo formato que el Excel original).
    Detecta las columnas por su CABECERA (mas robusto al orden cambiante).
    Filtra igual que el Excel: solo filas CONFIRMADA + ACCESO A VÍA=SI."""
    data = request.get_json(silent=True) or {}
    texto = (data.get('texto') or '').strip()
    if not texto:
        return jsonify({'error': 'No se recibió texto'}), 400

    # Normalizar separadores: tabs -> tab único; detectar si viene del Excel (tab)
    # Si no hay tabs, asumimos que son múltiples espacios y los convertimos a tabs
    if '\t' not in texto and '  ' in texto:
        texto = re.sub(r' {2,}', '\t', texto)

    lineas = [ln for ln in texto.split('\n') if ln.strip()]
    if not lineas:
        return jsonify({'error': 'Texto vacío'}), 400

    # Leer con csv dialect excel-tab (maneja comillas correctamente)
    reader = csv.reader(_stdio_io.StringIO(texto), dialect='excel-tab', skipinitialspace=False)

    # Mapeo de cabeceras reconocidas -> campo interno
    # (claves normalizadas en MAYUS sin tildes ni espacios dobles)
    MAPA_CABECERAS = {
        'ESTADO': 'estado',
        'SPCO QUE VALIDA': 'spco',
        'SEGUIMIENTO (NOMBRE OPCO)': 'operador',
        'SEGUIMIENTO (NOMBRE OPC)': 'operador',
        'ORDEN DE TRABAJO': 'orden_trabajo',
        'EMPRESA CONTRATISTA': 'empresa',
        'ACCESO A VÍA': 'acceso',
        'ACCESO A VIA': 'acceso',
        'DESDE': 'desde',
        'HASTA': 'hasta',
        'RESPONSABLE DE TRABAJOS': 'responsable',
        'CELULAR Y TETRA EXT': 'celular_tetra',
        # Comentario del permiso (con prioridad para el más descriptivo)
        'REQUERIMIENTOS ADICIONALES / COMENTARIOS': 'comentario',
        'REQUERIMIENTOS ADICIONALE': 'comentario',
        'REQUERIMENTOS ADICIONALES / COMENTARIOS': 'comentario',
        'COMENTARIOS DEL PCO': 'comentario_pco',
    }

    def norm(s):
        # Normaliza: MAYUS, sin tildes, espacios colapsados
        s = (s or '').strip().upper()
        s = s.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U').replace('Ñ', 'N')
        s = re.sub(r'\s+', ' ', s)
        return s

    # Primera fila NO vacia = cabecera. Buscarla.
    fila_iter = iter(reader)
    cabecera_idx = {}
    cab_encontrada = False
    for fila in fila_iter:
        if not fila or not any(c.strip() for c in fila):
            continue
        norm_fila = [norm(c) for c in fila]
        # Comprobar si contiene al menos 'ESTADO' y 'ACCESO A VIA' -> es cabecera
        if 'ESTADO' in norm_fila or 'ESTADO' in [n.split('(')[0].strip() for n in norm_fila]:
            for i, c in enumerate(norm_fila):
                if c in MAPA_CABECERAS:
                    cabecera_idx[MAPA_CABECERAS[c]] = i
            cab_encontrada = True
            break
        # Puede que la primera fila ya sea datos (sin cabecera) -> caer al fallback fijo
        else:
            # Reinsertar esta fila como primera fila de datos usando mapeo fijo
            datos_fijos = [_parsear_fila_fija(fila)]
            cab_encontrada = False
            break

    if not cab_encontrada and not cabecera_idx:
        # Fallback: usar mapeo posicional fijo (texto sin cabecera, igual que antes)
        # Releer todo desde el principio
        reader = csv.reader(_stdio_io.StringIO(texto), dialect='excel-tab', skipinitialspace=False)
        datos = []
        for fila in reader:
            parsed = _parsear_fila_fija(fila)
            if parsed:
                datos.append(parsed)
    else:
        # Mapeo por cabecera: leer resto de filas
        campos_requeridos = ('estado', 'acceso')
        if not all(c in cabecera_idx for c in campos_requeridos):
            return jsonify({'error': 'Faltan cabeceras obligatorias (ESTADO, ACCESO A VÍA). Cabeceras detectadas: ' + ', '.join(cabecera_idx.keys())}), 400

        datos = []
        for fila in fila_iter:
            if not fila or not any(c.strip() for c in fila):
                continue
            estado_val = fila[cabecera_idx['estado']].strip().upper() if 'estado' in cabecera_idx and len(fila) > cabecera_idx['estado'] else ''
            acceso_val = fila[cabecera_idx['acceso']].strip().upper() if 'acceso' in cabecera_idx and len(fila) > cabecera_idx['acceso'] else ''

            if estado_val != 'CONFIRMADA':
                continue
            if acceso_val != 'SI':
                continue

            def getcampo(campo):
                idx = cabecera_idx.get(campo)
                if idx is None or idx >= len(fila):
                    return ''
                return str(fila[idx]).strip()

            celular = getcampo('celular_tetra')
            tetra = extract_tetra(celular)

            comentario_permiso = ''
            if 'comentario' in cabecera_idx:
                comentario_permiso = getcampo('comentario').strip()
            if not comentario_permiso and 'comentario_pco' in cabecera_idx:
                comentario_permiso = getcampo('comentario_pco').strip()
            if comentario_permiso in ('\xa0', 'N/A', 'n/a'):
                comentario_permiso = ''

            datos.append({
                'spco': getcampo('spco'),
                'operador': getcampo('operador'),
                'desde': getcampo('desde'),
                'hasta': getcampo('hasta'),
                'responsable': getcampo('responsable'),
                'tetra': tetra,
                'orden_trabajo': getcampo('orden_trabajo'),
                'empresa': getcampo('empresa'),
                'comentario': comentario_permiso,
            })

    if not datos:
        return jsonify({'error': 'No se encontraron filas CONFIRMADAS con ACCESO=SI en el texto pegado'}), 400

    # Guardar en sesión (igual que importar_excel)
    prev = session.get('importados_pendientes', [])
    prev.extend(datos)
    session['importados_pendientes'] = prev

    return jsonify({'datos': datos, 'total': len(datos)})


@app.route('/limpiar_importados', methods=['POST'])
def limpiar_importados():
    """Limpia toda la lista de importados pendientes de la sesion."""
    session['importados_pendientes'] = []
    return jsonify({'success': True, 'limpiados': True})


def _parsear_fila_fija(fila):
    """Fallback: mapeo posicional fijo cuando NO hay cabecera reconocida.
    Mantiene compatibilidad con el mapeo anterior."""
    if not fila or len(fila) < 5:
        return None
    col0 = (fila[0] or '').strip().upper()
    if col0 in ('N°', 'NRO', 'N.', 'N', 'ITEM', '#'):
        return None
    estado = (fila[1] if len(fila) > 1 else '') or ''
    acceso = (fila[13] if len(fila) > 13 else '') or ''
    if str(estado).strip().upper() != 'CONFIRMADA':
        return None
    if str(acceso).strip().upper() != 'SI':
        return None
    spco = (fila[2] if len(fila) > 2 else '') or ''
    operador = (fila[4] if len(fila) > 4 else '') or ''
    orden_trabajo = (fila[7] if len(fila) > 7 else '') or ''
    empresa = (fila[10] if len(fila) > 10 else '') or ''
    desde = (fila[16] if len(fila) > 16 else '') or ''
    hasta = (fila[17] if len(fila) > 17 else '') or ''
    responsable = (fila[19] if len(fila) > 19 else '') or ''
    celular_tetra = (fila[20] if len(fila) > 20 else '') or ''
    tetra = extract_tetra(celular_tetra)
    # REQUERIMIENTOS ADICIONALES / COMENTARIOS era col 23 (index 22) en el Excel
    comentario = (fila[22] if len(fila) > 22 else '') or ''
    if str(comentario).strip() in ('\xa0', 'N/A', 'n/a'):
        comentario = ''
    return {
        'spco': str(spco).strip(),
        'operador': str(operador).strip(),
        'desde': str(desde).strip(),
        'hasta': str(hasta).strip(),
        'responsable': str(responsable).strip(),
        'tetra': tetra,
        'orden_trabajo': str(orden_trabajo).strip(),
        'empresa': str(empresa).strip(),
        'comentario': str(comentario).strip(),
    }


@app.route('/confirmar_importado', methods=['POST'])
def confirmar_importado():
    data = request.get_json() or {}
    idx = data.get('_idx', -1)

    # Tomar datos completos desde la sesión
    pendientes = session.get('importados_pendientes', [])
    item = pendientes[idx] if 0 <= idx < len(pendientes) else data

    turno = session.get('turno', 'Turno 1')
    operador = item.get('operador', session.get('operador_turno', 'Importado'))
    spco = item.get('spco', session.get('spco_turno', 'Importado'))
    empresa = item.get('empresa', '')
    orden_trabajo = item.get('orden_trabajo', '')
    responsable = item.get('responsable', '')
    desde = item.get('desde', '')
    hasta = item.get('hasta', '')
    zona = f"{desde} -> {hasta}" if desde and hasta else (desde or hasta)
    tetra = item.get('tetra', '')
    comentario = item.get('comentario', '')

    hora_inicio_str = data.get('hora_inicio', '')
    hora_inicio = datetime.strptime(hora_inicio_str, "%H:%M:%S").time() if hora_inicio_str else None

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO seguimiento_vias 
        (turno, operador_turno, spco_turno, empresa, orden_trabajo, responsable, ubicacion_zona, num_personas, tetra, fecha, hora_inicio, estado, comentario)
        VALUES (%s, %s, %s, %s, %s, %s, %s, 0, %s, CURRENT_DATE, COALESCE(%s, CURRENT_TIME), 'En Vía', %s)
        RETURNING id;
    """, (turno, operador, spco, empresa, orden_trabajo, responsable, zona, tetra, hora_inicio, comentario))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()

    # Remover de la sesión
    pendientes = session.get('importados_pendientes', [])
    if 0 <= idx < len(pendientes):
        pendientes.pop(idx)
        session['importados_pendientes'] = pendientes

    return jsonify({'success': True, 'id': new_id, 'empresa': empresa, 'comentario': comentario})


@app.route('/revertir_importado/<int:id>')
def revertir_importado(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT operador_turno, spco_turno, empresa, orden_trabajo, responsable, ubicacion_zona, tetra, comentario
        FROM seguimiento_vias WHERE id = %s;
    """, (id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({'error': 'No encontrado'}), 404

    operador, spco, empresa, orden_trabajo, responsable, zona, tetra, comentario = row

    cur.execute("DELETE FROM seguimiento_vias WHERE id = %s;", (id,))
    conn.commit()
    cur.close()
    conn.close()

    # Separar zona en desde -> hasta
    desde, hasta = (zona.split(' -> ') + ['', ''])[:2] if ' -> ' in zona else (zona, '')

    item = {
        'spco': spco or '',
        'operador': operador or '',
        'desde': desde,
        'hasta': hasta,
        'responsable': responsable or '',
        'tetra': tetra or '',
        'orden_trabajo': orden_trabajo or '',
        'empresa': empresa or '',
        'comentario': comentario or '',
    }

    pendientes = session.get('importados_pendientes', [])
    pendientes.append(item)
    session['importados_pendientes'] = pendientes

    return jsonify({'success': True, 'item': item})


# ==========================================================================
# ADMIN: Panel oculto protegido por PIN
# ==========================================================================
ADMIN_PIN = os.environ.get("ADMIN_PIN", "admin123")  # PIN por defecto local; en produccion usar variable de entorno ADMIN_PIN

def admin_verificar_pin(pin):
    """Compara PIN con el configurado en variables de entorno."""
    if not pin:
        return False
    return pin.strip() == ADMIN_PIN


def _clasificar_bloque(hora_inicio):
    """Devuelve 'dia', 'noche', o 'sin_hora' segun el objeto time recibido."""
    if hora_inicio is None:
        return 'sin_hora'
    try:
        h = hora_inicio
        # 05:00 inclusive -> 17:00 exclusive = dia
        from datetime import time as _time
        if h >= _time(5, 0) and h < _time(17, 0):
            return 'dia'
        return 'noche'
    except Exception:
        return 'sin_hora'


@app.route('/estadisticas')
def estadisticas():
    """Dashboard de estadisticas construido a partir de los registros ya
    almacenados (los mismos que ve el Historial), agregados en Python.

    Se hace UNA sola consulta SQL (identica en filtros a la del historial)
    y todos los conteos se calculan en Python. As evitamos problemas por
    casts de tipo time, NULLs, etc. que hacian que SQL perdiera algunos rows.

    Filtros aceptados (mismos nombres que el historial, para reenviarlos):
      - fecha_inicio / fecha_fin (rango de fechas)
      - turno: 'dia' (5-17), 'noche' (17-5), o '' (ambos)
    """
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    turno_filtro = request.args.get('turno', '')

    # Construccion del WHERE identica en esencia al historial (sin bloque_horario,
    # porque aqui usamos el 'turno' simplificado a dia/noche/ambos).
    where_clauses = []
    params = []
    if fecha_inicio:
        where_clauses.append("fecha >= %s"); params.append(fecha_inicio)
    if fecha_fin:
        where_clauses.append("fecha <= %s"); params.append(fecha_fin)
    if turno_filtro == 'dia':
        # Rango simple dentro del mismo dia: 05:00 a 17:00 (igual que historial bloque_horario=dia)
        where_clauses.append("hora_inicio >= '05:00'::time AND hora_inicio <= '17:00'::time")
    elif turno_filtro == 'noche':
        # Rango que cruza medianoche: 17:00 a 05:00 (igual que historial bloque_horario=noche)
        where_clauses.append("(hora_inicio >= '17:00'::time OR hora_inicio <= '05:00'::time)")
    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    # Acumuladores en Python
    total_reg = 0
    dias_turno_labels = []; dias_dia_data = []; dias_noche_data = []
    dias_labels = []; dias_data = []
    bloques = {'Dia (05-17)': 0, 'Noche (17-05)': 0, 'Sin hora': 0}
    empresas_contador = {}
    respon_contador = {}
    estados = {'En Vía': 0, 'Liberado': 0}
    dias_semana_contador = {}
    zonas_contador = {}
    duraciones_min = []
    fechas_existentes = []
    total_sin_filtro = 0

    conn = None; cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # CONSULTA PRINCIPAL: trae todos los registros que cumplen el filtro
        # (identico al historial). Indices:
        # 0 id, 1 fecha, 2 turno, 3 empresa, 4 orden_trabajo, 5 responsable,
        # 6 ubicacion_zona, 7 num_personas, 8 tetra, 9 hora_inicio, 10 hora_fin,
        # 11 estado, 12 usa_vehiculo, 13 tipo_vehiculo, 14 codigo_vehiculo,
        # 15 conductor_vehiculo, 16 operador_turno, 17 spco_turno, 18 comentario
        cur.execute(f"""
            SELECT id, fecha, turno, empresa, orden_trabajo, responsable, ubicacion_zona,
                   num_personas, tetra, hora_inicio, hora_fin, estado,
                   usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo,
                   operador_turno, spco_turno, comentario
            FROM seguimiento_vias{where_sql}
            ORDER BY fecha ASC, hora_inicio ASC;
        """, tuple(params))
        registros = cur.fetchall()

        total_reg = len(registros)
        # Mapeo dia de la semana
        semana_map_idx = {'MON':'Lun','TUE':'Mar','WED':'Mié','THU':'Jue','FRI':'Vie','SAT':'Sáb','SUN':'Dom'}
        orden_semana = ['Lun','Mar','Mié','Jue','Vie','Sáb','Dom']
        # Agregaciones por fecha (preservando orden cronologico)
        por_dia_acum = {}  # fecha --> {'dia':int,'noche':int,'sin_hora':int,'total':int}
        for r in registros:
            fecha = str(r[1]) if r[1] is not None else 'S/F'
            empresa = (r[3] or '').strip()
            responsable = (r[5] or '').strip()
            estado = r[11] or ''
            ubicacion = (r[6] or '').strip()
            hora_inicio = r[9]

            # Bloque dia/noche segun hora_inicio
            bloque = _clasificar_bloque(hora_inicio)
            bloque_key = {'dia':'Dia (05-17)', 'noche':'Noche (17-05)', 'sin_hora':'Sin hora'}[bloque]
            bloques[bloque_key] = bloques.get(bloque_key, 0) + 1

            # Agregaciones por empresa/responsable/estado/zona
            if empresa:
                empresas_contador[empresa] = empresas_contador.get(empresa, 0) + 1
            if responsable:
                # Juntar nombres sin normalizar excesivamente: tomamos la primera linea para agrupar mejor
                resp_clean = responsable.split('\n')[0].strip()
                if resp_clean:
                    respon_contador[resp_clean] = respon_contador.get(resp_clean, 0) + 1
            if estado in estados:
                estados[estado] += 1
            else:
                estados[estado] = estados.get(estado, 0) + 1
            if ubicacion:
                # Normalizar zonas agrupando por el primer componente antes de ' -> '
                zona_clean = ubicacion.split(' -> ')[0].split(',')[0].strip()
                if zona_clean:
                    zonas_contador[zona_clean] = zonas_contador.get(zona_clean, 0) + 1

            # Duracion (si inicio y fin presentes)
            if r[9] is not None and r[10] is not None:
                try:
                    # Devuelve minutos
                    if hasattr(r[10], 'timestamp') and hasattr(r[9], 'timestamp'):
                        # son time, no datetime; usar differencia manual
                        pass
                    # Trabajamos con objetos time; construir datetime dummy
                    from datetime import datetime as _dt
                    base = _dt(2000, 1, 1)
                    hi = r[9]; hf = r[10]
                    if hasattr(hi, 'hour'):
                        dt0 = base.replace(hour=hi.hour, minute=hi.minute, second=hi.second or 0)
                        dt1 = base.replace(hour=hf.hour, minute=hf.minute, second=hf.second or 0)
                        if dt1 < dt0:
                            dt1 = dt1.replace(day=base.day + 1)
                        seconds = (dt1 - dt0).total_seconds()
                        if seconds > 0:
                            duraciones_min.append(seconds / 60.0)
                except Exception:
                    pass

            # Dia de la semana (name en ingles, mapeado a Lun-Dom)
            if r[1] is not None:
                try:
                    dia_en = r[1].strftime('%a').upper()  # ej 'MON'
                    dia_es = semana_map_idx.get(dia_en, dia_en.title())
                    dias_semana_contador[dia_es] = dias_semana_contador.get(dia_es, 0) + 1
                except Exception:
                    pass

            # Agregacion por fecha
            if fecha not in por_dia_acum:
                por_dia_acum[fecha] = {'dia':0, 'noche':0, 'sin_hora':0, 'total':0}
            por_dia_acum[fecha]['total'] += 1
            por_dia_acum[fecha][bloque] += 1

        for fecha in sorted(por_dia_acum.keys()):
            dias_turno_labels.append(fecha)
            dias_dia_data.append(por_dia_acum[fecha]['dia'])
            dias_noche_data.append(por_dia_acum[fecha]['noche'] + por_dia_acum[fecha].get('sin_hora', 0))
            dias_labels.append(fecha)
            dias_data.append(por_dia_acum[fecha]['total'])

        # Top 5 empresas
        empresas_sorted = sorted(empresas_contador.items(), key=lambda x: x[1], reverse=True)[:5]
        empresas_labels = [e[0] for e in empresas_sorted]
        empresas_data = [int(e[1]) for e in empresas_sorted]

        # Top 5 responsables
        respon_sorted = sorted(respon_contador.items(), key=lambda x: x[1], reverse=True)[:5]
        respon_labels = [e[0] for e in respon_sorted]
        respon_data = [int(e[1]) for e in respon_sorted]

        # Top 5 zonas
        zonas_sorted = sorted(zonas_contador.items(), key=lambda x: x[1], reverse=True)[:5]
        zonas_labels = [e[0] for e in zonas_sorted]
        zonas_data = [int(e[1]) for e in zonas_sorted]

        # Dia de la semana en orden Lun-Dom
        dias_semana_labels = [d for d in orden_semana if d in dias_semana_contador]
        dias_semana_data = [dias_semana_contador.get(d, 0) for d in dias_semana_labels]

        # Totales
        total_dia = bloques['Dia (05-17)']
        total_noche = bloques['Noche (17-05)']
        total_archivados = total_reg  # en este contexto, equivalente
        promedio_duracion_min = (sum(duraciones_min) / len(duraciones_min)) if duraciones_min else 0.0

        # Diagnostico: fechas existentes (sin filtros) y total sin filtro
        cur.execute("SELECT fecha, COUNT(*) FROM seguimiento_vias GROUP BY fecha ORDER BY fecha DESC LIMIT 30;")
        for r in cur.fetchall():
            fechas_existentes.append((str(r[0]), int(r[1])))
        cur.execute("SELECT COUNT(*) FROM seguimiento_vias;")
        row = cur.fetchone()
        total_sin_filtro = int(row[0]) if row and row[0] is not None else 0

    except (OperationalError, DatabaseError) as e:
        app.logger.error(f"estadisticas DB error: {e}")
        # Todo se queda en valores vacios
    finally:
        if cur: cur.close()
        if conn: conn.close()

    return render_template('estadisticas.html',
        dias_labels=dias_labels, dias_data=dias_data,
        dias_turno_labels=dias_turno_labels, dias_dia_data=dias_dia_data, dias_noche_data=dias_noche_data,
        bloques=bloques,
        total_dia=total_dia, total_noche=total_noche,
        empresas_labels=empresas_labels, empresas_data=empresas_data,
        respon_labels=respon_labels, respon_data=respon_data,
        estados=estados,
        total_reg=total_reg,
        total_archivados=total_archivados,
        total_lib=int(estados.get('Liberado', 0)),
        total_via=int(estados.get('En Vía', 0)),
        dias_semana_labels=dias_semana_labels, dias_semana_data=dias_semana_data,
        promedio_duracion_min=promedio_duracion_min,
        zonas_labels=zonas_labels, zonas_data=zonas_data,
        fechas_existentes=fechas_existentes,
        total_sin_filtro=total_sin_filtro,
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, turno_filtro=turno_filtro)


@app.route('/estadisticas_debug')
def estadisticas_debug():
    """ENDPOINT TEMPORAL DE DEBUG.
    Devuelve JSON con todas las consultas SQL que ejecuta /estadisticas
    para los mismos filtros, junto con sus resultados. Sirve para diagnosticar
    por qué las estadísticas no muestran datos cuando el historial sí los tiene.
    NO requiere autorización; se debe eliminar en cuanto se termine el debug."""
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    turno_filtro = request.args.get('turno', '')

    where_clauses = []
    params = []
    if fecha_inicio:
        where_clauses.append("fecha >= %s"); params.append(fecha_inicio)
    if fecha_fin:
        where_clauses.append("fecha <= %s"); params.append(fecha_fin)
    if turno_filtro == 'dia':
        where_clauses.append("hora_inicio IS NOT NULL AND hora_inicio::time >= '05:00'::time AND hora_inicio::time < '17:00'::time")
    elif turno_filtro == 'noche':
        where_clauses.append("hora_inicio IS NOT NULL AND (hora_inicio::time >= '17:00'::time OR hora_inicio::time < '05:00'::time)")
    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    out = {
        'filtros': {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'turno': turno_filtro},
        'where_sql': where_sql or '(sin WHERE)',
        'params': params,
        'consultas': []
    }

    conn = None; cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # 0) Distintos estados reales que hay en la BD
        q = "SELECT estado, COUNT(*) FROM seguimiento_vias GROUP BY estado ORDER BY COUNT(*) DESC;"
        cur.execute(q)
        out['consultas'].append({
            'nombre': 'estados_unicos',
            'sql': q,
            'resultado': cur.fetchall()
        })

        # 1) Todos los registros con sus fechas/horas (muestra primeros 50 ordenados por fecha desc)
        q = "SELECT id, fecha, hora_inicio, hora_fin, estado, turno, empresa, responsable FROM seguimiento_vias ORDER BY fecha DESC, hora_inicio DESC NULLS LAST LIMIT 50;"
        cur.execute(q)
        out['consultas'].append({
            'nombre': 'todos_50_recientes',
            'sql': q,
            'resultado': cur.fetchall()
        })

        # 2) Trabajos por fecha con el WHERE aplicado
        q1 = f"SELECT fecha, COUNT(*) AS total FROM seguimiento_vias{where_sql} GROUP BY fecha ORDER BY fecha ASC;"
        cur.execute(q1, tuple(params))
        out['consultas'].append({
            'nombre': 'por_dia_con_filtro',
            'sql': q1,
            'params': params,
            'resultado': cur.fetchall()
        })

        # 3) Bloques dia/noche con el WHERE aplicado
        q2 = f"""SELECT CASE WHEN hora_inicio IS NOT NULL AND hora_inicio::time >= '05:00'::time AND hora_inicio::time < '17:00'::time THEN 'Dia (05-17)' ELSE 'Noche (17-05)' END AS bloque, COUNT(*) AS total
                 FROM seguimiento_vias{where_sql} GROUP BY bloque;"""
        cur.execute(q2, tuple(params))
        out['consultas'].append({
            'nombre': 'bloques_dia_noche_con_filtro',
            'sql': q2,
            'params': params,
            'resultado': cur.fetchall()
        })

        # 4) Estados con el WHERE aplicado (lo que usa total_reg)
        q3 = f"SELECT estado, COUNT(*) FROM seguimiento_vias{where_sql} GROUP BY estado;"
        cur.execute(q3, tuple(params))
        out['consultas'].append({
            'nombre': 'estados_con_filtro',
            'sql': q3,
            'params': params,
            'resultado': cur.fetchall()
        })

        # 5) Total con filtro (lo que usa total_filtrado)
        q4 = f"SELECT COUNT(*) FROM seguimiento_vias{where_sql};"
        cur.execute(q4, tuple(params))
        out['consultas'].append({
            'nombre': 'total_con_filtro',
            'sql': q4,
            'params': params,
            'resultado': cur.fetchone()
        })

        # 6) Total sin filtro
        cur.execute("SELECT COUNT(*) FROM seguimiento_vias;")
        out['consultas'].append({
            'nombre': 'total_sin_filtro',
            'sql': 'SELECT COUNT(*) FROM seguimiento_vias;',
            'resultado': cur.fetchone()
        })

        # 7) Fechas disponibles (done el diagnóstico)
        cur.execute("SELECT fecha, COUNT(*) AS total FROM seguimiento_vias GROUP BY fecha ORDER BY fecha DESC LIMIT 30;")
        out['consultas'].append({
            'nombre': 'fechas_disponibles',
            'sql': 'SELECT fecha, COUNT(*) AS total FROM seguimiento_vias GROUP BY fecha ORDER BY fecha DESC LIMIT 30;',
            'resultado': cur.fetchall()
        })

        # 8) Verificación: cuantos tienen hora_inicio = NULL por fecha
        cur.execute("SELECT fecha, COUNT(*) AS con_hora_null FROM seguimiento_vias WHERE hora_inicio IS NULL GROUP BY fecha ORDER BY fecha DESC LIMIT 30;")
        out['consultas'].append({
            'nombre': 'registros_sin_hora_inicio_por_fecha',
            'sql': 'SELECT fecha, COUNT(*) AS con_hora_null FROM seguimiento_vias WHERE hora_inicio IS NULL GROUP BY fecha ORDER BY fecha DESC LIMIT 30;',
            'resultado': cur.fetchall()
        })

        # 9) Zona horaria efectiva de la conexión
        cur.execute("SHOW TIME ZONE;")
        out['consultas'].append({
            'nombre': 'timezone_actual',
            'sql': "SHOW TIME ZONE;",
            'resultado': cur.fetchone()
        })

        # 10) CURRENT_DATE y CURRENT_TIME actual en la sesión
        cur.execute("SELECT CURRENT_DATE, CURRENT_TIME, NOW();")
        out['consultas'].append({
            'nombre': 'current_date_time',
            'sql': "SELECT CURRENT_DATE, CURRENT_TIME, NOW();",
            'resultado': cur.fetchone()
        })

    except Exception as e:
        out['ERROR'] = f"{type(e).__name__}: {e}"
    finally:
        if cur: cur.close()
        if conn: conn.close()

    import json as _json
    def _default(o):
        if isinstance(o, (datetime, date)):
            return o.isoformat()
        if hasattr(o, 'isoformat'):
            return o.isoformat()
        return str(o)
    return _json.dumps(out, ensure_ascii=False, indent=2, default=_default), 200, {'Content-Type': 'application/json; charset=utf-8'}


@app.route('/admin')
def admin_panel():
    """Panel admin: muestra TODOS los registros historicos (no solo hoy).
    Requiere PIN via query (?pin=XXX) o via session['admin_pin_ok'].
    Si no hay PIN o es incorrecto, devuelve 404 (página no encontrada)
    para no revelar la existencia de la ruta."""
    pin = request.args.get('pin', '').strip()
    pin_session = session.get('admin_pin_ok', False)

    if not pin and not pin_session:
        # Sin PIN -> devuelve 404 para aparentar que la pagina no existe
        return render_template('admin.html', autorizado=False), 404
    if pin and not admin_verificar_pin(pin):
        # PIN incorrecto -> 404 (no revelar)
        return render_template('admin.html', autorizado=False), 404
    if pin and admin_verificar_pin(pin):
        # PIN correcto -> guardar autorizacion en sesion para futuras visitas
        session['admin_pin_ok'] = True
        session['admin_pin_valor'] = pin

    # Cargar TODOS los registros
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, fecha, turno, empresa, orden_trabajo, responsable, ubicacion_zona,
                   num_personas, tetra, hora_inicio, hora_fin, estado,
                   usa_vehiculo, tipo_vehiculo, codigo_vehiculo, conductor_vehiculo,
                   operador_turno, spco_turno, archivado, comentario
            FROM seguimiento_vias
            ORDER BY fecha DESC, id DESC;
        """, )
        registros = cur.fetchall()
    except (OperationalError, DatabaseError) as e:
        app.logger.error(f"admin_panel DB error: {e}")
        registros = []
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    return render_template('admin.html', autorizado=True, registros=registros)


@app.route('/comentar/<int:id>', methods=['POST'])
def comentar(id):
    """Actualiza solo el campo comentario de un registro (edicion inline)."""
    data = request.get_json(silent=True) or {}
    comentario = (data.get('comentario') or '').strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE seguimiento_vias SET comentario = %s WHERE id = %s;", (comentario, id))
        conn.commit()
    except (OperationalError, DatabaseError) as e:
        app.logger.error(f"comentar DB error: {e}")
        return jsonify({'error': 'Error de base de datos'}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()
    return jsonify({'success': True, 'comentario': comentario})


@app.route('/admin/eliminar_todos', methods=['POST'])
def admin_eliminar_todos():
    if not session.get('admin_pin_ok'):
        data = request.get_json(silent=True) or {}
        pin = (data.get('pin') or '').strip()
        if not admin_verificar_pin(pin):
            return jsonify({'error': 'No autorizado'}), 403

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM seguimiento_vias;")
        total = cur.fetchone()[0]
        cur.execute("DELETE FROM seguimiento_vias;")
        conn.commit()
        app.logger.warning(f"admin_eliminar_todos: {total} registros eliminados de la base de datos")
    except (OperationalError, DatabaseError) as e:
        app.logger.error(f"admin_eliminar_todos DB error: {e}")
        return jsonify({'error': 'Error de base de datos'}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    return jsonify({'success': True, 'eliminados': total})


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_pin_ok', None)
    session.pop('admin_pin_valor', None)
    return jsonify({'success': True})


@app.route('/admin/eliminar/<int:id>', methods=['POST'])
def admin_eliminar(id):
    # Verificar autorizacion en sesion primero
    if not session.get('admin_pin_ok'):
        # Permitir override con PIN en el body (por si caduca la sesion)
        data = request.get_json(silent=True) or {}
        pin = (data.get('pin') or '').strip()
        if not admin_verificar_pin(pin):
            return jsonify({'error': 'No autorizado'}), 403

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM seguimiento_vias WHERE id = %s;", (id,))
        conn.commit()
    except (OperationalError, DatabaseError) as e:
        app.logger.error(f"admin_eliminar DB error: {e}")
        return jsonify({'error': 'Error de base de datos'}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    return jsonify({'success': True, 'id': id})


if __name__ == '__main__':
    app.run(debug=True, port=5000)